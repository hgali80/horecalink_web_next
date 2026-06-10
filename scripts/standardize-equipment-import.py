from __future__ import annotations

import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SOURCE = Path(
    r"C:\Users\hasan\OneDrive\Belgeler\HORECL~1\1-YENI~1\horecalink_urunleri_tam_liste_v3.xlsx"
)
DEFAULT_OUTPUT = Path(
    r"C:\Users\hasan\OneDrive\Belgeler\HORECL~1\1-YENI~1\horecalink_urunleri_tam_liste_v4_equipment_import.xlsx"
)


def build_canonical_maps(rows, header_index):
    categories = OrderedDict()
    subcategories = {}

    for row in rows:
      group_key = str(row[header_index["groupKey"]] or "").strip()
      if group_key != "equipment":
        continue

      category_key = str(row[header_index["categoryKey"]] or "").strip()
      category_label = str(row[header_index["category"]] or "").strip()
      subcategory_key = str(row[header_index["subcategoryKey"]] or "").strip()
      subcategory_label = str(row[header_index["subcategory"]] or "").strip()

      if category_key and category_key not in categories:
        categories[category_key] = category_label

      if subcategory_key and subcategory_key not in subcategories:
        subcategories[subcategory_key] = {
          "categoryKey": category_key,
          "categoryLabel": category_label,
          "subcategoryLabel": subcategory_label,
        }

    return categories, subcategories


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    wb = load_workbook(source)
    ws = wb["Urun_Sablonu"]

    header = [cell.value for cell in ws[4]]
    header_index = {name: idx for idx, name in enumerate(header)}
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    categories, subcategories = build_canonical_maps(rows, header_index)

    for row_idx in range(5, ws.max_row + 1):
        group_key = str(ws.cell(row=row_idx, column=header_index["groupKey"] + 1).value or "").strip()
        if group_key != "equipment":
            continue

        subcategory_key = str(
            ws.cell(row=row_idx, column=header_index["subcategoryKey"] + 1).value or ""
        ).strip()
        category_key = str(
            ws.cell(row=row_idx, column=header_index["categoryKey"] + 1).value or ""
        ).strip()

        if subcategory_key in subcategories:
            canonical = subcategories[subcategory_key]
        else:
            canonical = {
                "categoryKey": category_key,
                "categoryLabel": categories.get(category_key, ""),
                "subcategoryLabel": str(
                    ws.cell(row=row_idx, column=header_index["subcategory"] + 1).value or ""
                ).strip(),
            }

        ws.cell(row=row_idx, column=header_index["group"] + 1, value="Ekipman")
        ws.cell(row=row_idx, column=header_index["groupKey"] + 1, value="equipment")
        ws.cell(
            row=row_idx,
            column=header_index["category"] + 1,
            value=canonical["categoryLabel"],
        )
        ws.cell(
            row=row_idx,
            column=header_index["categoryKey"] + 1,
            value=canonical["categoryKey"],
        )
        ws.cell(
            row=row_idx,
            column=header_index["subcategory"] + 1,
            value=canonical["subcategoryLabel"],
        )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(
        {
            "source": str(source),
            "output": str(output),
            "categoryCount": len(categories),
            "subcategoryCount": len(subcategories),
        }
    )


if __name__ == "__main__":
    main()
